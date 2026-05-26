import polars as pl
from hereutil import here
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time

# -------------------------------------------------------
# MARC21 Country Code → Human-Readable Name Mapping
# -------------------------------------------------------
COUNTRY_MAP = {
    # Core Western European
    "gw": "Germany", "de": "Germany", "fr": "France", "it": "Italy",
    "sp": "Spain", "es": "Spain", "pt": "Portugal", "PT": "Portugal",
    "enk": "England", "gb": "Great Britain", "xxk": "United Kingdom",
    "stk": "Scotland", "ie": "Ireland",
    "nl": "Netherlands", "ne": "Netherlands",
    "be": "Belgium", "sz": "Switzerland", "ch": "Switzerland",
    "au": "Austria", 
    # Scandinavia
    "sw": "Sweden", "dk": "Denmark", "fi": "Finland",
    "no": "Norway", "no#": "Norway",
    # Central & Eastern Europe
    "pl": "Poland", "hu": "Hungary", "cz": "Czech Republic",
    "xr": "Czech Republic", "xn": "Netherlands",
    "ru": "Russia", "er": "Estonia", "eu": "Estonia",
    # Americas & Asia
    "us": "United States", "xxu": "United States",
    "ca": "Canada", "ja": "Japan", "cc": "China", "ii": "India",
    # Extended XA- / XG- / XD- prefixed codes (regional catalog standards)
    "XA-DE": "Germany", "XA-DXDE": "Germany (East)",
    "XA-FR": "France", "XA-GB": "Great Britain",
    "XA-IT": "Italy", "XA-ES": "Spain", "XA-AT": "Austria",
    "XA-CH": "Switzerland", "XA-NL": "Netherlands",
    "XA-SE": "Sweden", "XA-PL": "Poland", "XA-DK": "Denmark",
    "XA-FI": "Finland", "XA-NO": "Norway", "XA-HU": "Hungary",
    "XA-BE": "Belgium", "XA-PT": "Portugal", "XA-CZ": "Czech Republic",
    "XG-DR": "Germany (East)", "XD-US": "United States",
    # Unknown / Undetermined / Junk
    "xx": "Unknown", "ZZ": "Unknown", "oo": "Unknown", "un": "Unknown",
    "|": "Unknown", "|||": "Unknown", "null": "Unknown", None: "Unknown",
}

# -------------------------------------------------------
# ISO 639-2 Language Code → Human-Readable Name Mapping
# -------------------------------------------------------
LANGUAGE_MAP = {
    "eng": "English", "ger": "German", "deu": "German",
    "fre": "French", "fra": "French", "lat": "Latin",
    "swe": "Swedish", "pol": "Polish", "ita": "Italian",
    "por": "Portuguese", "spa": "Spanish",
    "dut": "Dutch", "nld": "Dutch", "ned": "Dutch", "dui": "Dutch",
    "rus": "Russian", "dan": "Danish",
    "fin": "Finnish", "hun": "Hungarian",
    "nor": "Norwegian", "nob": "Norwegian (Bokmål)", "nno": "Norwegian (Nynorsk)",
    "cze": "Czech", "ces": "Czech",
    "est": "Estonian",
    "grc": "Ancient Greek", "gre": "Greek", "ell": "Greek",
    "heb": "Hebrew", "ara": "Arabic", "tur": "Turkish",
    "jpn": "Japanese", "chi": "Chinese", "zho": "Chinese",
    "kor": "Korean", "hin": "Hindi",
    "zxx": "No Linguistic Content", "mul": "Multiple Languages",
    "mmm": "Unknown (Legacy)", "und": "Undetermined",
    "...": "Unknown", "null": "Unknown", None: "Unknown",
}

def resolve_country(code):
    if code is None:
        return "Unknown"
    return COUNTRY_MAP.get(str(code).strip(), str(code).upper())

def resolve_language(code):
    if code is None:
        return "Unknown"
    code_str = str(code).strip()
    # Filter out dirty data: years appearing as language codes
    if code_str.isdigit():
        return "Unknown (Data Error)"
    return LANGUAGE_MAP.get(code_str, code_str.lower())


def create_excel():
    print("Loading CSV baseline data...")
    csv_path = here("global_baseline.csv")
    df = pl.read_csv(csv_path)
    
    print("Calculating aggregations...")
    # 1. Decade Aggregation
    df_decade = (
        df
        .with_columns(((pl.col("year_of_publication") // 10) * 10).alias("decade"))
        .group_by("decade")
        .agg(pl.col("total_publications").sum().alias("count"))
        .sort("decade")
    )
    
    # 2. Country Aggregation (with human-readable names)
    df_country_raw = (
        df
        .group_by("country_of_publication")
        .agg(pl.col("total_publications").sum().alias("count"))
        .sort("count", descending=True)
    )
    # Map codes to names and re-aggregate (since multiple codes map to same country)
    country_rows = {}
    for row in df_country_raw.iter_rows():
        name = resolve_country(row[0])
        country_rows[name] = country_rows.get(name, 0) + row[1]
    df_country_list = sorted(country_rows.items(), key=lambda x: x[1], reverse=True)[:15]
    
    # 3. Language Aggregation (with human-readable names)
    df_language_raw = (
        df
        .group_by("primary_language_code")
        .agg(pl.col("total_publications").sum().alias("count"))
        .sort("count", descending=True)
    )
    lang_rows = {}
    for row in df_language_raw.iter_rows():
        name = resolve_language(row[0])
        lang_rows[name] = lang_rows.get(name, 0) + row[1]
    df_language_list = sorted(lang_rows.items(), key=lambda x: x[1], reverse=True)[:15]
    
    # 4. Yearly Timeline
    df_yearly = (
        df
        .group_by("year_of_publication")
        .agg(pl.col("total_publications").sum().alias("count"))
        .sort("year_of_publication")
    )
    
    print("Creating Excel workbook...")
    wb = openpyxl.Workbook()
    
    # Define styles
    navy_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    gray_fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    title_font = Font(name="Segoe UI", size=16, bold=True, color="1F4E79")
    subtitle_font = Font(name="Segoe UI", size=10, italic=True, color="595959")
    section_font = Font(name="Segoe UI", size=12, bold=True, color="1F4E79")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Segoe UI", size=11, bold=True)
    regular_font = Font(name="Segoe UI", size=11)
    footnote_font = Font(name="Segoe UI", size=9, italic=True, color="808080")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    double_bottom_border = Border(
        top=Side(style='thin', color='1F4E79'),
        bottom=Side(style='double', color='1F4E79')
    )

    # ----------------------------------------------------
    # TAB 1: EXECUTIVE SUMMARY
    # ----------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    ws_summary.views.sheetView[0].showGridLines = False
    
    # Title Block — merged across entire width
    ws_summary.merge_cells("A1:H1")
    ws_summary["A1"] = "DHH26 Global Book Production Baseline"
    ws_summary["A1"].font = title_font
    ws_summary["A1"].alignment = Alignment(vertical="center")
    
    ws_summary.merge_cells("A2:H2")
    ws_summary["A2"] = "A historical overview of publication volumes (1401–2026), compiled from 99,026,673 catalog records across multiple European national bibliographies."
    ws_summary["A2"].font = subtitle_font
    ws_summary["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws_summary.row_dimensions[2].height = 30
    
    # ---- Decade Table ----
    ws_summary["A4"] = "Publication Volume by Decade"
    ws_summary["A4"].font = section_font
    
    headers_decade = ["Decade", "Publications"]
    for col_idx, h in enumerate(headers_decade, start=1):
        cell = ws_summary.cell(row=5, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center" if col_idx == 1 else "right")
    
    row_num = 6
    for row in df_decade.iter_rows():
        ws_summary.cell(row=row_num, column=1, value=f"{row[0]}s").alignment = Alignment(horizontal="center")
        ws_summary.cell(row=row_num, column=2, value=row[1]).number_format = "#,##0"
        
        fill = gray_fill if row_num % 2 == 0 else white_fill
        for c in range(1, 3):
            cell = ws_summary.cell(row=row_num, column=c)
            cell.font = regular_font
            cell.fill = fill
            cell.border = thin_border
        row_num += 1
        
    # Decade Total Row
    ws_summary.cell(row=row_num, column=1, value="Total").font = bold_font
    ws_summary.cell(row=row_num, column=1).alignment = Alignment(horizontal="center")
    total_decade_cell = ws_summary.cell(row=row_num, column=2, value=f"=SUM(B6:B{row_num-1})")
    total_decade_cell.font = bold_font
    total_decade_cell.number_format = "#,##0"
    total_decade_cell.border = double_bottom_border
    ws_summary.cell(row=row_num, column=1).border = Border(top=Side(style='thin', color='1F4E79'))

    # ---- Top Countries Table ----
    ws_summary["D4"] = "Top 15 Publishing Countries"
    ws_summary["D4"].font = section_font
    
    headers_country = ["Country", "Publications"]
    for col_idx, h in enumerate(headers_country, start=4):
        cell = ws_summary.cell(row=5, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center" if col_idx == 4 else "right")
        
    row_num_c = 6
    for name, count in df_country_list:
        ws_summary.cell(row=row_num_c, column=4, value=name).alignment = Alignment(horizontal="left", indent=1)
        ws_summary.cell(row=row_num_c, column=5, value=count).number_format = "#,##0"
        
        fill = gray_fill if row_num_c % 2 == 0 else white_fill
        for c in range(4, 6):
            cell = ws_summary.cell(row=row_num_c, column=c)
            cell.font = regular_font
            cell.fill = fill
            cell.border = thin_border
        row_num_c += 1
        
    ws_summary.cell(row=row_num_c, column=4, value="Total (Top 15)").font = bold_font
    total_country_cell = ws_summary.cell(row=row_num_c, column=5, value=f"=SUM(E6:E{row_num_c-1})")
    total_country_cell.font = bold_font
    total_country_cell.number_format = "#,##0"
    total_country_cell.border = double_bottom_border
    ws_summary.cell(row=row_num_c, column=4).border = Border(top=Side(style='thin', color='1F4E79'))

    # ---- Top Languages Table ----
    ws_summary["G4"] = "Top 15 Publishing Languages"
    ws_summary["G4"].font = section_font
    
    headers_lang = ["Language", "Publications"]
    for col_idx, h in enumerate(headers_lang, start=7):
        cell = ws_summary.cell(row=5, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center" if col_idx == 7 else "right")
        
    row_num_l = 6
    for name, count in df_language_list:
        ws_summary.cell(row=row_num_l, column=7, value=name).alignment = Alignment(horizontal="left", indent=1)
        ws_summary.cell(row=row_num_l, column=8, value=count).number_format = "#,##0"
        
        fill = gray_fill if row_num_l % 2 == 0 else white_fill
        for c in range(7, 9):
            cell = ws_summary.cell(row=row_num_l, column=c)
            cell.font = regular_font
            cell.fill = fill
            cell.border = thin_border
        row_num_l += 1
        
    ws_summary.cell(row=row_num_l, column=7, value="Total (Top 15)").font = bold_font
    total_lang_cell = ws_summary.cell(row=row_num_l, column=8, value=f"=SUM(H6:H{row_num_l-1})")
    total_lang_cell.font = bold_font
    total_lang_cell.number_format = "#,##0"
    total_lang_cell.border = double_bottom_border
    ws_summary.cell(row=row_num_l, column=7).border = Border(top=Side(style='thin', color='1F4E79'))

    # ---- Metadata Footer ----
    footer_row = max(row_num, row_num_c, row_num_l) + 3
    ws_summary.merge_cells(f"A{footer_row}:H{footer_row}")
    ws_summary.cell(row=footer_row, column=1, value="Data Source: GizmoSQL FlightSQL Database (grpc+tls://gizmosql-dhh26.2.rahtiapp.fi) — DHH26 Helsinki Digital Humanities Hackathon 2026").font = footnote_font
    ws_summary.merge_cells(f"A{footer_row+1}:H{footer_row+1}")
    ws_summary.cell(row=footer_row+1, column=1, value="Extraction Date: 20 May 2026. Country codes mapped from MARC21 standard. Language codes mapped from ISO 639-2.").font = footnote_font

    # Column widths for summary
    ws_summary.column_dimensions["A"].width = 16
    ws_summary.column_dimensions["B"].width = 18
    ws_summary.column_dimensions["C"].width = 4
    ws_summary.column_dimensions["D"].width = 22
    ws_summary.column_dimensions["E"].width = 18
    ws_summary.column_dimensions["F"].width = 4
    ws_summary.column_dimensions["G"].width = 26
    ws_summary.column_dimensions["H"].width = 18

    # ----------------------------------------------------
    # TAB 2: YEARLY TIMELINE
    # ----------------------------------------------------
    ws_yearly = wb.create_sheet(title="Yearly Timeline")
    ws_yearly.views.sheetView[0].showGridLines = False
    
    ws_yearly.merge_cells("A1:B1")
    ws_yearly["A1"] = "Year-by-Year Publication Timeline"
    ws_yearly["A1"].font = title_font
    ws_yearly["A1"].alignment = Alignment(vertical="center")
    
    headers_y = ["Year of Publication", "Publications"]
    for col_idx, h in enumerate(headers_y, start=1):
        cell = ws_yearly.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center" if col_idx == 1 else "right")
    
    # Freeze panes below header
    ws_yearly.freeze_panes = "A4"
        
    row_num_y = 4
    for row in df_yearly.iter_rows():
        ws_yearly.cell(row=row_num_y, column=1, value=row[0]).alignment = Alignment(horizontal="center")
        ws_yearly.cell(row=row_num_y, column=2, value=row[1]).number_format = "#,##0"
        
        fill = gray_fill if row_num_y % 2 == 0 else white_fill
        for c in range(1, 3):
            cell = ws_yearly.cell(row=row_num_y, column=c)
            cell.font = regular_font
            cell.fill = fill
            cell.border = thin_border
        row_num_y += 1
        
    ws_yearly.cell(row=row_num_y, column=1, value="Grand Total").font = bold_font
    ws_yearly.cell(row=row_num_y, column=1).alignment = Alignment(horizontal="center")
    total_y_cell = ws_yearly.cell(row=row_num_y, column=2, value=f"=SUM(B4:B{row_num_y-1})")
    total_y_cell.font = bold_font
    total_y_cell.number_format = "#,##0"
    total_y_cell.border = double_bottom_border
    ws_yearly.cell(row=row_num_y, column=1).border = Border(top=Side(style='thin', color='1F4E79'))
    
    ws_yearly.column_dimensions["A"].width = 22
    ws_yearly.column_dimensions["B"].width = 18

    # ----------------------------------------------------
    # TAB 3: RAW GROUPED DATA
    # ----------------------------------------------------
    print("Writing raw data tab... (This writes 435k rows, might take a moment)")
    ws_raw = wb.create_sheet(title="Raw Data (Grouped)")
    ws_raw.views.sheetView[0].showGridLines = True
    
    headers_raw = ["Year", "Country", "Language", "Publications"]
    for col_idx, h in enumerate(headers_raw, start=1):
        cell = ws_raw.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center" if col_idx < 4 else "right")
    
    # Freeze panes below header
    ws_raw.freeze_panes = "A2"
        
    # Write data rows with resolved names
    row_num_r = 2
    for row in df.iter_rows():
        ws_raw.cell(row=row_num_r, column=1, value=row[0])
        ws_raw.cell(row=row_num_r, column=2, value=resolve_country(row[1]))
        ws_raw.cell(row=row_num_r, column=3, value=resolve_language(row[2]))
        pub_cell = ws_raw.cell(row=row_num_r, column=4, value=row[3])
        pub_cell.number_format = "#,##0"
        row_num_r += 1
    
    ws_raw.column_dimensions["A"].width = 12
    ws_raw.column_dimensions["B"].width = 22
    ws_raw.column_dimensions["C"].width = 26
    ws_raw.column_dimensions["D"].width = 18
    
    # Apply AutoFilter to raw data
    ws_raw.auto_filter.ref = f"A1:D{row_num_r - 1}"
            
    output_excel = here("global_baseline.xlsx")
    print(f"Saving Excel workbook to {output_excel}...")
    wb.save(output_excel)
    print("Excel workbook created successfully!")

if __name__ == "__main__":
    start = time.time()
    create_excel()
    print(f"Completed in {time.time() - start:.2f} seconds")
