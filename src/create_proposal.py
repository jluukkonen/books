import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time

def create_proposal():
    wb = openpyxl.Workbook()

    # ── Colour palette ──
    navy = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    dark_teal = PatternFill(start_color="0D4F4F", end_color="0D4F4F", fill_type="solid")
    light_gray = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
    white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Member colours (matching Miro board)
    joonas_fill = PatternFill(start_color="DAEEF7", end_color="DAEEF7", fill_type="solid")   # light blue
    udita_fill = PatternFill(start_color="E8F5C9", end_color="E8F5C9", fill_type="solid")    # light green
    duong_fill = PatternFill(start_color="E8DEF8", end_color="E8DEF8", fill_type="solid")    # lavender
    sophia_fill = PatternFill(start_color="FCDADA", end_color="FCDADA", fill_type="solid")   # pink
    annalisa_fill = PatternFill(start_color="C9DAF8", end_color="C9DAF8", fill_type="solid") # medium blue
    orange_fill = PatternFill(start_color="FDE9D0", end_color="FDE9D0", fill_type="solid")   # orange
    
    # Fonts
    title_font = Font(name="Segoe UI", size=20, bold=True, color="1F4E79")
    subtitle_font = Font(name="Segoe UI", size=12, italic=True, color="595959")
    section_font = Font(name="Segoe UI", size=14, bold=True, color="1F4E79")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Segoe UI", size=11, bold=True)
    regular_font = Font(name="Segoe UI", size=11)
    small_font = Font(name="Segoe UI", size=10)
    small_italic = Font(name="Segoe UI", size=10, italic=True, color="595959")
    quote_font = Font(name="Segoe UI", size=12, italic=True, color="1F4E79")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # =====================================================
    # TAB 1: OVERVIEW
    # =====================================================
    ws = wb.active
    ws.title = "Project Overview"
    ws.views.sheetView[0].showGridLines = False
    
    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "Disruption & Resilience in European Knowledge Production"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells("A2:H2")
    ws["A2"] = "DHH26 Group Research Proposal — Helsinki Digital Humanities Hackathon 2026"
    ws["A2"].font = subtitle_font
    
    ws.merge_cells("A4:H4")
    ws["A4"] = "\"European book production has been punctuated by massive disruptions — wars, plagues, revolutions, censorship regimes —"
    ws["A4"].font = quote_font
    ws["A4"].alignment = Alignment(wrap_text=True)
    
    ws.merge_cells("A5:H5")
    ws["A5"] = "yet it always recovered and often transformed in the process. Are these transformations random, or do they follow predictable patterns?\""
    ws["A5"].font = quote_font
    ws["A5"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[4].height = 25
    ws.row_dimensions[5].height = 25
    
    # ── Core Question ──
    ws.merge_cells("A7:H7")
    ws["A7"] = "Core Research Question"
    ws["A7"].font = section_font
    
    ws.merge_cells("A8:H9")
    cell = ws["A8"]
    cell.value = (
        "For each major historical disruption (war, plague, revolution), we measure six analytical dimensions — "
        "volume & geography, marginal voices, network structure, concept diffusion, cultural heritage revival, "
        "and genre/language shifts — across three time windows (before, during, after). "
        "We then compare across disruptions to see if a common pattern of transformation emerges."
    )
    cell.font = regular_font
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[8].height = 25
    ws.row_dimensions[9].height = 25
    
    # ── Disruption Events Table ──
    ws.merge_cells("A11:H11")
    ws["A11"] = "Candidate Disruption Events (Select 3)"
    ws["A11"].font = section_font
    
    event_headers = ["Event", "Period", "Type", "Why It's Interesting"]
    col_widths_events = [4, 5, 5, 8]  # spans
    event_col_starts = [1, 2, 3, 4]
    
    for col_idx, h in enumerate(event_headers, start=1):
        cell = ws.cell(row=12, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = dark_teal
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    # Merge columns E-H for the "Why" column
    ws.merge_cells("D12:H12")
    ws.cell(row=12, column=4).font = header_font
    ws.cell(row=12, column=4).fill = dark_teal
    ws.cell(row=12, column=4).alignment = Alignment(horizontal="center")
    for c in range(5, 9):
        ws.cell(row=12, column=c).fill = dark_teal
    
    events = [
        ("⭐ The Reformation", "1517–1555", "Religious", "Explosion of pamphlets; Latin → vernacular shift"),
        ("⭐ Thirty Years' War", "1618–1648", "Military", "Devastated German printing; did other centers rise?"),
        ("Great Plague of London", "1665–1666", "Epidemic", "Did English publishing shift topics toward medicine?"),
        ("⭐ French Revolution", "1789–1799", "Political", "Censorship collapse → explosion → Napoleonic censorship"),
        ("Napoleonic Wars", "1803–1815", "Military", "Redrew the map; triggered nationalism & heritage revivals"),
        ("Revolutions of 1848", "1848", "Political", "Simultaneous across Europe; ideal for cross-country comparison"),
        ("⭐ World War I", "1914–1918", "Total War", "Language shifts, propaganda, collapse of empires"),
        ("World War II", "1939–1945", "Total War", "Book burning, exile publishers, post-war reconstruction"),
    ]
    
    for i, (event, period, etype, why) in enumerate(events):
        row = 13 + i
        ws.cell(row=row, column=1, value=event).font = bold_font if "⭐" in event else regular_font
        ws.cell(row=row, column=2, value=period).font = regular_font
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3, value=etype).font = regular_font
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        ws.merge_cells(f"D{row}:H{row}")
        ws.cell(row=row, column=4, value=why).font = small_font
        ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True)
        
        fill = light_gray if i % 2 == 0 else white
        for c in range(1, 9):
            ws.cell(row=row, column=c).fill = fill
            ws.cell(row=row, column=c).border = thin_border
    
    ws.merge_cells(f"A22:H22")
    ws.cell(row=22, column=1, value="⭐ = Recommended selection: covers three centuries and three different disruption types.").font = small_italic
    
    # Column widths
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14

    # =====================================================
    # TAB 2: TEAM ROLES
    # =====================================================
    ws2 = wb.create_sheet(title="Team Roles & Dimensions")
    ws2.views.sheetView[0].showGridLines = False
    
    ws2.merge_cells("A1:F1")
    ws2["A1"] = "Six Analytical Dimensions — One Per Team Member"
    ws2["A1"].font = title_font
    ws2["A1"].alignment = Alignment(vertical="center")
    ws2.row_dimensions[1].height = 40
    
    ws2.merge_cells("A2:F2")
    ws2["A2"] = "Each member owns one analytical \"lens\". For each disruption event, all six lenses are applied to the same three time windows (before / during / after)."
    ws2["A2"].font = subtitle_font
    ws2["A2"].alignment = Alignment(wrap_text=True)
    ws2.row_dimensions[2].height = 35
    
    role_headers = ["Member", "Dimension", "Key Question", "What They Measure", "Tools / Data", "Output"]
    for col_idx, h in enumerate(role_headers, start=1):
        cell = ws2.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = navy
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    
    roles = [
        (
            "Sophia", sophia_fill,
            "The Macro View\n(Volume & Geography)",
            "What happens to the overall system?",
            "Total publication counts;\nCountry-level market share shifts;\nRecovery time (years to pre-disruption levels)",
            "global_baseline.xlsx;\nYearly Timeline tab",
            "Volume charts;\nGeographic heatmaps;\nRecovery curves"
        ),
        (
            "Udita", udita_fill,
            "The Social View\n(Marginal Voices)",
            "Who benefits and who suffers?",
            "Proportion of female / marginal authors;\nNew vs. established author ratios;\nDoes upheaval open doors or consolidate hierarchies?",
            "dhh_search_tool.py;\nAuthor metadata queries",
            "Before/after proportion charts;\nCase study narratives"
        ),
        (
            "Duong", duong_fill,
            "The Structural View\n(Network Reconfiguration)",
            "How does publishing infrastructure reshape?",
            "Author-publisher bipartite networks;\nClustering coefficient, degree distribution;\nCentralization vs. fragmentation",
            "network_exporter.py;\nGephi / NetworkX",
            "Network visualisations;\nStructural metrics table"
        ),
        (
            "Joonas", joonas_fill,
            "The Semantic View\n(Concept Diffusion)",
            "What ideas accelerate or stall?",
            "Geographic spread of key concepts;\nFirst-appearance mapping;\nDiffusion speed before vs. after disruption",
            "dhh_search_tool.py;\nglobal_baseline.xlsx",
            "Concept flow maps;\nDiffusion timelines"
        ),
        (
            "Annalisa", annalisa_fill,
            "The Heritage View\n(Cultural Memory)",
            "Do disruptions trigger a 'return to roots'?",
            "Publication frequency of medieval/classical texts;\nCorrelation with nationalist movements;\nRediscovery waves",
            "dhh_search_tool.py;\nCorpus identification",
            "Medieval text frequency charts;\nNationalism correlation"
        ),
        (
            "Member 6", orange_fill,
            "The Genre & Language View",
            "Does the intellectual diet change?",
            "Genre proportions (religious, political, scientific);\nLanguage proportions (Latin vs. vernacular);\nTranslation pattern shifts",
            "dhh_search_tool.py;\nglobal_baseline.xlsx;\nGenre keyword heuristics",
            "Genre pie charts over time;\nLanguage shift timelines"
        ),
    ]
    
    for i, (name, fill, dimension, question, measures, tools, output) in enumerate(roles):
        row = 5 + i
        ws2.cell(row=row, column=1, value=name).font = bold_font
        ws2.cell(row=row, column=1).fill = fill
        ws2.cell(row=row, column=2, value=dimension).font = bold_font
        ws2.cell(row=row, column=2).fill = fill
        ws2.cell(row=row, column=3, value=question).font = regular_font
        ws2.cell(row=row, column=4, value=measures).font = small_font
        ws2.cell(row=row, column=5, value=tools).font = small_font
        ws2.cell(row=row, column=6, value=output).font = small_font
        
        for c in range(1, 7):
            cell = ws2.cell(row=row, column=c)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        ws2.row_dimensions[row].height = 75
    
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 28
    ws2.column_dimensions["D"].width = 32
    ws2.column_dimensions["E"].width = 26
    ws2.column_dimensions["F"].width = 24

    # =====================================================
    # TAB 3: ANALYTICAL FRAMEWORK
    # =====================================================
    ws3 = wb.create_sheet(title="Analytical Framework")
    ws3.views.sheetView[0].showGridLines = False
    
    ws3.merge_cells("A1:G1")
    ws3["A1"] = "Shared Analytical Framework"
    ws3["A1"].font = title_font
    ws3.row_dimensions[1].height = 40
    
    ws3.merge_cells("A2:G2")
    ws3["A2"] = "For each disruption event, all six dimensions are measured across three time windows."
    ws3["A2"].font = subtitle_font
    ws3.row_dimensions[2].height = 25
    
    # Time windows diagram
    ws3.merge_cells("A4:G4")
    ws3["A4"] = "Time Windows (Example: Thirty Years' War, 1618–1648)"
    ws3["A4"].font = section_font
    
    pre_fill = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")
    during_fill = PatternFill(start_color="F8CECC", end_color="F8CECC", fill_type="solid")
    post_fill = PatternFill(start_color="DAE8FC", end_color="DAE8FC", fill_type="solid")
    
    windows = [
        ("B6", "C6", "PRE-DISRUPTION", "1588–1617", "(30 years before)", pre_fill),
        ("D6", "D6", "DURING", "1618–1648", "(the event itself)", during_fill),
        ("E6", "F6", "POST-DISRUPTION", "1649–1678", "(30 years after)", post_fill),
    ]
    
    for start, end, label, years, note, fill in windows:
        ws3.merge_cells(f"{start}:{end}")
        cell = ws3[start]
        cell.value = f"{label}\n{years}\n{note}"
        cell.font = Font(name="Segoe UI", size=12, bold=True, color="333333")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style='medium', color='333333'),
            right=Side(style='medium', color='333333'),
            top=Side(style='medium', color='333333'),
            bottom=Side(style='medium', color='333333')
        )
    ws3.row_dimensions[6].height = 65

    # Also fill merged cells
    for c_letter in ["C"]:
        ws3[f"{c_letter}6"].fill = pre_fill
        ws3[f"{c_letter}6"].border = Border(left=Side(style='medium', color='333333'), right=Side(style='medium', color='333333'), top=Side(style='medium', color='333333'), bottom=Side(style='medium', color='333333'))
    for c_letter in ["F"]:
        ws3[f"{c_letter}6"].fill = post_fill
        ws3[f"{c_letter}6"].border = Border(left=Side(style='medium', color='333333'), right=Side(style='medium', color='333333'), top=Side(style='medium', color='333333'), bottom=Side(style='medium', color='333333'))
    
    # Measurement matrix
    ws3.merge_cells("A9:G9")
    ws3["A9"] = "Measurement Matrix (What We Measure in Each Window)"
    ws3["A9"].font = section_font
    
    matrix_headers = ["Dimension", "Member", "Pre-Disruption", "During", "Post-Disruption", "Comparison Metric"]
    for col_idx, h in enumerate(matrix_headers, start=1):
        cell = ws3.cell(row=10, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = navy
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    
    matrix_rows = [
        ("Volume & Geography", "Sophia", "Baseline volume\n& country shares", "Volume drop?\nGeographic shift?", "Recovery speed?\nNew centers?", "Years to recovery;\n% share change"),
        ("Marginal Voices", "Udita", "% female/marginal\nauthors", "Does % rise\nor fall?", "Permanent gain\nor reversion?", "Δ proportion;\nNew entrants count"),
        ("Network Structure", "Duong", "Network density,\nclustering, hubs", "Fragmentation\nor consolidation?", "New structure\nor restoration?", "Δ clustering;\nΔ centralization"),
        ("Concept Diffusion", "Joonas", "Concept presence\nby country", "Acceleration\nor stalling?", "New geographies\nreached?", "Diffusion speed;\nNew countries reached"),
        ("Cultural Heritage", "Annalisa", "Medieval text\nfrequency", "Heritage interest\nduring crisis?", "Post-crisis\nheritage revival?", "Spike detection;\nCorrelation w/ nationalism"),
        ("Genre & Language", "Member 6", "Genre & language\nproportions", "Topic flight?\nVernacular shift?", "New genre\nequilibrium?", "Δ genre shares;\nΔ language ratios"),
    ]
    
    member_fills = [sophia_fill, udita_fill, duong_fill, joonas_fill, annalisa_fill, orange_fill]
    
    for i, (dim, member, pre, during, post, metric) in enumerate(matrix_rows):
        row = 11 + i
        ws3.cell(row=row, column=1, value=dim).font = bold_font
        ws3.cell(row=row, column=1).fill = member_fills[i]
        ws3.cell(row=row, column=2, value=member).font = bold_font
        ws3.cell(row=row, column=2).fill = member_fills[i]
        ws3.cell(row=row, column=3, value=pre).font = small_font
        ws3.cell(row=row, column=3).fill = pre_fill
        ws3.cell(row=row, column=4, value=during).font = small_font
        ws3.cell(row=row, column=4).fill = during_fill
        ws3.cell(row=row, column=5, value=post).font = small_font
        ws3.cell(row=row, column=5).fill = post_fill
        ws3.cell(row=row, column=6, value=metric).font = small_font
        
        for c in range(1, 7):
            cell = ws3.cell(row=row, column=c)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        
        ws3.row_dimensions[row].height = 50
    
    # Paper structure
    ws3.merge_cells("A19:G19")
    ws3["A19"] = "Proposed Paper Structure"
    ws3["A19"].font = section_font
    
    paper_sections = [
        ("1.", "Introduction", "The question & why it matters"),
        ("2.", "Data & Methods", "99M-record database, baseline normalization, shared time windows, six analytical dimensions"),
        ("3.", "Case Study 1", "Thirty Years' War (1618–1648) — all six dimensions applied"),
        ("4.", "Case Study 2", "French Revolution (1789–1799) — all six dimensions applied"),
        ("5.", "Case Study 3", "World War I (1914–1918) — all six dimensions applied"),
        ("6.", "Comparative Discussion", "Do disruptions follow a common pattern across centuries and shock types?"),
        ("7.", "Conclusion", "What we learned about how knowledge production responds to crisis"),
    ]
    
    for i, (num, section, desc) in enumerate(paper_sections):
        row = 20 + i
        ws3.cell(row=row, column=1, value=num).font = bold_font
        ws3.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws3.merge_cells(f"B{row}:C{row}")
        ws3.cell(row=row, column=2, value=section).font = bold_font
        ws3.merge_cells(f"D{row}:G{row}")
        ws3.cell(row=row, column=4, value=desc).font = regular_font
        ws3.cell(row=row, column=4).alignment = Alignment(wrap_text=True)
        
        fill = light_gray if i % 2 == 0 else white
        for c in range(1, 8):
            ws3.cell(row=row, column=c).fill = fill
            ws3.cell(row=row, column=c).border = thin_border
    
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 16
    ws3.column_dimensions["C"].width = 18
    ws3.column_dimensions["D"].width = 18
    ws3.column_dimensions["E"].width = 18
    ws3.column_dimensions["F"].width = 22
    ws3.column_dimensions["G"].width = 14

    # =====================================================
    # TAB 4: PRACTICAL NOTES
    # =====================================================
    ws4 = wb.create_sheet(title="Practical Notes")
    ws4.views.sheetView[0].showGridLines = False
    
    ws4.merge_cells("A1:E1")
    ws4["A1"] = "Practical Considerations"
    ws4["A1"].font = title_font
    ws4.row_dimensions[1].height = 40
    
    # Strengths
    ws4.merge_cells("A3:E3")
    ws4["A3"] = "Why This Approach Works"
    ws4["A3"].font = section_font
    
    strengths = [
        ("Clean division of labor", "Each member owns one dimension end-to-end. No overlapping work."),
        ("Shared baseline already built", "The global_baseline.xlsx (99M records) is ready and provides normalization for all members."),
        ("Tools are ready", "dhh_search_tool.py (concept search), network_exporter.py (network extraction), and the baseline cover all six dimensions."),
        ("Scales to available time", "If time is short: do 1 disruption event. More time: do 3. The method is identical."),
        ("Academically grounded", "Builds on Tolonen et al. (2021) and Marjanen et al. (2025), both already identified on the Miro board."),
        ("Visually compelling", "The final output includes maps, network graphs, timelines, and cross-dimensional comparisons."),
    ]
    
    for i, (title, desc) in enumerate(strengths):
        row = 4 + i
        ws4.cell(row=row, column=1, value="✅").font = regular_font
        ws4.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws4.merge_cells(f"B{row}:C{row}")
        ws4.cell(row=row, column=2, value=title).font = bold_font
        ws4.merge_cells(f"D{row}:E{row}")
        ws4.cell(row=row, column=4, value=desc).font = regular_font
        ws4.cell(row=row, column=4).alignment = Alignment(wrap_text=True)
        
        fill = light_gray if i % 2 == 0 else white
        for c in range(1, 6):
            ws4.cell(row=row, column=c).fill = fill
            ws4.cell(row=row, column=c).border = thin_border
    
    # Timeline
    ws4.merge_cells("A12:E12")
    ws4["A12"] = "Suggested Work Timeline"
    ws4["A12"].font = section_font
    
    timeline_headers = ["Phase", "When", "What", "Who"]
    for col_idx, h in enumerate(timeline_headers, start=1):
        cell = ws4.cell(row=13, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = navy
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    ws4.merge_cells("C13:D13")
    ws4.cell(row=13, column=3).font = header_font
    ws4.cell(row=13, column=3).fill = navy
    ws4.cell(row=13, column=4).fill = navy
    
    timeline = [
        ("1. Align", "Day 1 (Today)", "Agree on 3 disruption events and confirm roles", "All"),
        ("2. Extract", "Day 1–2", "Each member runs their queries for pre/during/post windows", "Individual"),
        ("3. Analyse", "Day 2–3", "Each member produces their dimension's charts and metrics", "Individual"),
        ("4. Synthesize", "Day 3–4", "Compare across disruptions; identify common patterns", "All"),
        ("5. Present", "Day 4–5", "Assemble final presentation / paper draft", "All"),
    ]
    
    for i, (phase, when, what, who) in enumerate(timeline):
        row = 14 + i
        ws4.cell(row=row, column=1, value=phase).font = bold_font
        ws4.cell(row=row, column=2, value=when).font = regular_font
        ws4.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        ws4.merge_cells(f"C{row}:D{row}")
        ws4.cell(row=row, column=3, value=what).font = regular_font
        ws4.cell(row=row, column=3).alignment = Alignment(wrap_text=True)
        ws4.cell(row=row, column=5, value=who).font = regular_font
        ws4.cell(row=row, column=5).alignment = Alignment(horizontal="center")
        
        fill = light_gray if i % 2 == 0 else white
        for c in range(1, 6):
            ws4.cell(row=row, column=c).fill = fill
            ws4.cell(row=row, column=c).border = thin_border
    
    ws4.column_dimensions["A"].width = 18
    ws4.column_dimensions["B"].width = 18
    ws4.column_dimensions["C"].width = 30
    ws4.column_dimensions["D"].width = 24
    ws4.column_dimensions["E"].width = 16
    
    # Save
    from hereutil import here
    output = here("group_proposal.xlsx")
    print(f"Saving to {output}...")
    wb.save(output)
    print("Done!")

if __name__ == "__main__":
    start = time.time()
    create_proposal()
    print(f"Completed in {time.time() - start:.2f} seconds")
